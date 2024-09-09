import openpyxl


workBook = openpyxl.load_workbook("ExcelFile.xlsx")

print(workBook.sheetnames) # print sheetname in the excel workbook as list

# for selecting particular sheets

# now we can work with sheet 1
worksheet1 = workBook['Sheet1']
print(worksheet1)


# we can use this to work with workshetet2
worksheet2 = workBook['Sheet2'] # case sensitive
print(worksheet2)

# for creating a newsheet
workBook.create_sheet('Sheet3')

# after doing all work save the workbook
workBook.save("ExcelFile.xlsx")
