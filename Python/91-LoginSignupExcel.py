# making login and signup page using python

# login page class to store email and password
class LoginPage:
    def __init__(self,email = "",password = ""):
        self.email = email
        self.password = password

    def setEmail(self,email) :
        self.email = email
    
    def setPassword(self,password):
        self.password = password

    def getEmail(self):
        return self.email
    
    def getPassword(self):
        return self.password
    
    def __str__(self) -> str:
        return f"Login Page"

# signup page class to store account information
class SignUpPage(LoginPage):
    def __init__(self,username = "",email = "",password = "",security = ""):
        LoginPage.__init__(self,email,password)
        self.username = username
        self.security = security

    def setUserName(self,username):
        self.username = username

    def getUserName(self):
        return self.username
    
    def setSecurity(self,secure):
        self.security = secure

    def getSecurity(self):
        return self.security
    
    def __str__(self) -> str:
        return f"SignUp Page"
    
# to get the file path
import os
# to deal with excel files
import openpyxl
# to deal with string formatting
import re

# for creating gile if not exist
if not os.path.exists("Page.xlsx"):
    wb  = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Username','Email',"Password","Security Questions"])
    wb.save("Page.xlsx")
else:
    wb = openpyxl.load_workbook('Page.xlsx')

# for accessing sheet
if wb.sheetnames:
    ws = wb[wb.sheetnames[0]]
else:
    wb.create_sheet("Sheet")
    ws = wb['Sheet']

# for checking the email format
def checkEmail():
    while True:
        user = input("Enter the email : ")
        email = re.search(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b",user)
        if email:
            return email.group()
        else:
            print("Please enter a valid email address.")

# function to add data in file
def addData(obj ):
    data = [obj.getUserName(),obj.getEmail(),obj.getPassword(),obj.getSecurity()]
    ws.append(data)
    wb.save('Page.xlsx')
    print('Account Created Succesfully.')


def choicescheck(txt):
    while True:
        userinput = input(txt)

        if userinput == 'Y' or userinput == 'y':
            return True
        elif userinput == 'n' or userinput == 'N':
            return False
        else:
            print("Invalid Input.")

def securityQuestions():
    questions = ["What your school name?","In which city do you live?","Whats your hometown?","Whats your nickname?"]

    string = str()
    rem = list()
    x = 0
    while x < 3:

        for y in rem:
            if y in questions:
                questions.remove(y)

        for Z in questions:
            print(Z)
            
            if choicescheck("Do you want to Skip this Question(y/n)?"):
                continue

            q = input()
            x += 1
            rem.append(Z)

            if string != "":
                string += ","
            string += Z + ":" + q  

    return string



# function to create account
def signup():
    username = input("Enter the username : ")
    found = True
    while found:
        email = checkEmail()

        lt = ws.iter_cols(min_col=2,max_col=2)
        found = False
        for e in lt:
            for v in e:
                if v.value == email:
                    found = True
                    print("Email Already Exists.")
                    break

    password = input("Enter the password : ")

    obj = SignUpPage(username,email,password,securityQuestions())
    addData(obj)

def forgotPassword(user):
    st = user.getSecurity().value
    st = st.split(",")
    print("You have to give At least two correct answers.")
    z = 0
    rem = list()
    while z < 2:
        for x in rem:
            if x in st:
                st.remove(x)

        for x in st:
            y = x.split(":")

            while True:
                userinput = input(y[0])

                if choicescheck("Do you want to skip this question(y/n)?"):
                    break
                
                if userinput == y[1]:
                    z += 1
                    rem.append(x)
                    break
                else:
                    print('Wrong Answer.')

                    
    newpass = input("Enter the new Password : ")
    user.getPassword().value = newpass
    wb.save('Page.xlsx')
        


# function to login into the account
def Login():
    email = checkEmail()
    lt = ws.iter_rows()
    user = SignUpPage() 
    found = False
    for u,e,p,s in lt:
        if e.value == email:
            user.setUserName(u)
            user.setEmail(e)
            user.setPassword(p)
            user.setSecurity(s)
            found = True
            break

    if found:
        while True:
            password = input("Enter the password : ")
            if password == user.getPassword().value:
                print('Login Succesfully.')
                AfterLogin(user)
                break
            else:
                print('Incorrect username or password.')
                if choicescheck("Do you want to reset Password.(y/n)?"):
                    forgotPassword(user)

    else:
        print("Username not found.")

# function to change the username
def changeUsername(user = SignUpPage()):
    username = input("Enter the new username : ")
    
    user.getUserName().value = username
    wb.save('Page.xlsx') 
    print("Username change Succesfully.")

#function to change the password
def changePassword(user = SignUpPage()):
    password = input("Enter the new password : ")
    
    user.getPassword().value = password
    wb.save('Page.xlsx') 
    print("Password Change Succesfully.")
    

def AfterLogin(user = SignUpPage()):
    while True:
        print("Username : ",user.getUserName().value)
        print("Email    : ",user.getEmail().value)
        print("Password : ",user.getPassword().value)

        print("1.Change Username")
        print("2.Change Password")
        print("3.Log Out")

        choice = input("Enter your choice : ")
        choice = int(choice)
        if choice == 1:
            changeUsername(user)
        elif choice == 2:
            changePassword(user)
        elif choice == 3:
            break
        else:
            print("Invalid Choice.")

#main function
def main():
    while True:
        print("1.Signup")
        print("2.Login")
        print("3.Exit")

        choice = input("Enter your choice : ")
        choice = int(choice)
        if choice == 1:
            signup()
        elif choice == 2:
            Login()
        elif choice == 3:
            break
        else:
            print("Invalid Choice.")

# main function running
main()