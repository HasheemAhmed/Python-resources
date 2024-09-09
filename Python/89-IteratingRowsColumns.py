import openpyxl

wb = openpyxl.load_workbook('ExcelFile.xlsx')

ws = wb['Sheet1']

# iter_rows() - to get the rows

row = ws.iter_rows(min_row=2) # min_row means start fron row 2

for x in row:
    for y in x:
        print(y.value)


# iter_cols() - to get the rows

row = ws.iter_rows(min_col=1,max_col=1) # min_row means start fron col 2

for x in row:
    for y in x:
        print(y.value)

    
# these two functions are used to get the values from the excel file
