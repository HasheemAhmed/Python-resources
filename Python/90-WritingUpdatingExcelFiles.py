import openpyxl 
import os

import openpyxl.workbook

if not os.path.exists("EF.xlsx"):
    wb = openpyxl.Workbook()
    wb.save('EF.xlsx')

wb = openpyxl.load_workbook("EF.xlsx")

wb.create_sheet("Info")

row = ['Name','Age']
ws = wb['Info']
ws.append(row)
wb.save('EF.xlsx')